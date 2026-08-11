"""Bulk import transactions from an Excel (.xlsx) spreadsheet."""

from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from models.account import Account
from models.category import Category
from models.envelope import Envelope
from models.investment import Investment
from models.transaction import Transaction
from services.transaction_service import TransactionValidationError, create_transaction


class ImportValidationError(ValueError):
    pass


TEMPLATE_HEADERS = [
    "date",
    "amount",
    "description",
    "type",
    "account",
    "to_account",
    "category",
    "paid_by",
    "payment_mode",
    "need_want",
    "envelope",
    "notes",
]

# Pre-filled data rows in the download (beyond header). User edits these.
TEMPLATE_PREFILL_ROWS = 40


def _joint_account() -> Account | None:
    return (
        Account.query.filter_by(name="Joint Account", is_active=True).first()
        or Account.query.filter_by(owner="joint", account_type="joint", is_active=True).first()
        or Account.query.filter_by(owner="joint", is_active=True).first()
    )


def build_template_xlsx() -> bytes:
    """Downloadable spreadsheet with dropdowns, Joint defaults, and auto envelope."""
    accounts = (
        Account.query.filter_by(is_active=True)
        .order_by(Account.sort_order, Account.name)
        .all()
    )
    categories = (
        Category.query.filter_by(is_active=True, parent_id=None)
        .order_by(Category.sort_order, Category.name)
        .all()
    )
    envelopes = (
        Envelope.query.filter_by(is_active=True)
        .order_by(Envelope.sort_order, Envelope.name)
        .all()
    )
    joint = _joint_account()
    joint_name = joint.name if joint else (accounts[0].name if accounts else "Joint Account")

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="F8FAFC")
    auto_fill = PatternFill("solid", fgColor="0F766E")
    auto_font = Font(color="ECFDF5", bold=True)

    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill

    # Mark auto columns in row 1 comments via second header hint row? Keep single header.
    # Prefill rows: date/amount/description blank; type/account/payment auto.
    today = date.today().isoformat()
    sample_cats = [
        c.name
        for c in categories
        if c.category_type == "expense"
        and c.name
        in {
            "Groceries",
            "Dining Out",
            "Movies & Entertainment",
            "Auto / Cab",
            "Fruits & Vegetables",
            "Personal Care & Household",
            "Utilities",
            "Fuel & Bike",
        }
    ]
    if not sample_cats:
        sample_cats = [c.name for c in categories if c.category_type == "expense"][:6]

    for r_idx in range(2, TEMPLATE_PREFILL_ROWS + 2):
        # date — leave blank except first few samples
        if r_idx <= 4 and sample_cats:
            sheet.cell(r_idx, 1, today)
        # amount / description blank for speed
        sheet.cell(r_idx, 4, "expense")  # type
        sheet.cell(r_idx, 5, joint_name)  # account → Joint
        sheet.cell(r_idx, 9, "upi")  # payment_mode
        sheet.cell(r_idx, 10, "need")  # need_want
        # envelope auto from category (column G)
        sheet.cell(
            r_idx,
            11,
            f'=IF(G{r_idx}="","",IFERROR(VLOOKUP(G{r_idx},Reference!$B$2:$I$500,8,FALSE),""))',
        )
        if r_idx - 2 < len(sample_cats):
            sheet.cell(r_idx, 7, sample_cats[r_idx - 2])
            sheet.cell(
                r_idx,
                3,
                f"{sample_cats[r_idx - 2]} — replace me",
            )
            sheet.cell(r_idx, 2, 500)

    for col in range(1, len(TEMPLATE_HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 16
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["G"].width = 26
    sheet.column_dimensions["K"].width = 16

    # Tint auto-filled columns
    for col in (4, 5, 9, 11):  # type, account, payment_mode, envelope
        cell = sheet.cell(1, col)
        cell.fill = auto_fill
        cell.font = auto_font

    # Reference lists
    ref = wb.create_sheet("Reference")
    ref_headers = [
        "Accounts",
        "Categories",
        "CategoryType",
        "Envelopes",
        "Types",
        "PaidBy",
        "PaymentMode",
        "NeedWant",
        "CategoryEnvelope",  # col I — VLOOKUP target for envelope auto
    ]
    for col, header in enumerate(ref_headers, start=1):
        cell = ref.cell(1, col, header)
        cell.font = Font(bold=True)

    for i, acc in enumerate(accounts, start=2):
        ref.cell(i, 1, acc.name)
    for i, cat in enumerate(categories, start=2):
        ref.cell(i, 2, cat.name)
        ref.cell(i, 3, cat.category_type)
        env_name = cat.envelope.name if cat.envelope else ""
        ref.cell(i, 9, env_name)  # CategoryEnvelope aligned with Categories row
    for i, env in enumerate(envelopes, start=2):
        ref.cell(i, 4, env.name)
    for i, t in enumerate(Transaction.TRANSACTION_TYPES, start=2):
        ref.cell(i, 5, t)
    for i, v in enumerate(Transaction.PAID_BY_CHOICES, start=2):
        ref.cell(i, 6, v)
    for i, v in enumerate(Transaction.PAYMENT_MODES, start=2):
        ref.cell(i, 7, v)
    for i, v in enumerate(Transaction.NEED_WANT_CHOICES, start=2):
        ref.cell(i, 8, v)

    for col in range(1, 10):
        ref.column_dimensions[get_column_letter(col)].width = 24

    def _ref_range(col_letter: str, count: int) -> str:
        end = max(count + 1, 2)
        return f"Reference!${col_letter}$2:${col_letter}${end}"

    validations = [
        ("D", _ref_range("E", len(Transaction.TRANSACTION_TYPES)), "Pick type"),
        ("E", _ref_range("A", max(len(accounts), 1)), "Pick account (default Joint)"),
        ("F", _ref_range("A", max(len(accounts), 1)), "Pick to_account"),
        ("G", _ref_range("B", max(len(categories), 1)), "Pick category"),
        ("H", _ref_range("F", len(Transaction.PAID_BY_CHOICES)), "Pick paid_by"),
        ("I", _ref_range("G", len(Transaction.PAYMENT_MODES)), "Pick payment_mode"),
        ("J", _ref_range("H", len(Transaction.NEED_WANT_CHOICES)), "Pick need_want"),
        ("K", _ref_range("D", max(len(envelopes), 1)), "Auto from category — override ok"),
    ]
    for col_letter, formula, prompt in validations:
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error="Choose a value from the dropdown list.",
            promptTitle="Select",
            prompt=prompt,
        )
        dv.add(f"{col_letter}2:{col_letter}1000")
        sheet.add_data_validation(dv)

    instructions = wb.create_sheet("Instructions", 1)
    lines = [
        "Finance OS — Fast Transaction Import",
        "",
        "Green columns are pre-filled / auto:",
        f"  • type → expense   • account → {joint_name}   • payment_mode → upi",
        "  • envelope → filled from category (Essentials / Lifestyle / Shopping…)",
        "",
        "Minimum you need per row:",
        "  1. date   2. amount   3. description   4. category (dropdown)",
        "",
        "Tips",
        "  • Dining Out / Movies & Entertainment → Lifestyle envelope (auto).",
        "  • Groceries / Rent / Utilities → Essentials (auto).",
        "  • Shopping / Furniture / Electronics → Shopping (auto).",
        "  • Override account when spend is from Suhel / Seema instead of Joint.",
        "  • Leave envelope blank on upload — app still maps from category.",
        "  • Transfers: set type=transfer, account=From, to_account=To.",
        "",
        "Re-download this template after adding accounts/categories so lists stay current.",
        "Save as .xlsx → Transactions → Import Excel → Preview → Confirm.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = instructions.cell(i, 1, line)
        cell.alignment = Alignment(wrap_text=False)
    instructions["A1"].font = Font(bold=True, size=14)
    instructions.column_dimensions["A"].width = 96

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def stage_upload(file_obj: BinaryIO, *, filename: str = "") -> str:
    """Save uploaded .xlsx to staging; return staging token."""
    name = (filename or "").lower()
    if name and not name.endswith(".xlsx"):
        raise ImportValidationError("Please upload an Excel .xlsx file (not .xls or CSV).")
    data = file_obj.read()
    if not data:
        raise ImportValidationError("The uploaded file is empty.")
    staging = _staging_dir()
    token = uuid.uuid4().hex
    path = staging / f"{token}.xlsx"
    path.write_bytes(data)
    return token


def staged_path(token: str) -> Path:
    if not token or not re.fullmatch(r"[a-f0-9]{32}", token):
        raise ImportValidationError("Invalid import session. Upload the file again.")
    path = _staging_dir() / f"{token}.xlsx"
    if not path.is_file():
        raise ImportValidationError("Import preview expired. Upload the file again.")
    return path


def clear_staged(token: str) -> None:
    try:
        path = staged_path(token)
    except ImportValidationError:
        return
    path.unlink(missing_ok=True)


def preview_xlsx(file_obj: BinaryIO, *, filename: str = "") -> dict[str, Any]:
    """Parse + validate without writing transactions. Stages file for confirm."""
    token = stage_upload(file_obj, filename=filename)
    try:
        with staged_path(token).open("rb") as fh:
            result = _process_xlsx(fh, dry_run=True, skip_duplicates=True)
    except Exception:
        clear_staged(token)
        raise
    result["staging_token"] = token
    result["dry_run"] = True
    return result


def import_xlsx(
    file_obj: BinaryIO | None = None,
    *,
    filename: str = "",
    staging_token: str | None = None,
    skip_duplicates: bool = True,
) -> dict[str, Any]:
    """
    Parse and import rows from an uploaded .xlsx file.

    Prefer staging_token from a prior preview. Valid rows are committed one-by-one.
    """
    cleanup_token = None
    if staging_token:
        path = staged_path(staging_token)
        cleanup_token = staging_token
        fh = path.open("rb")
    else:
        if file_obj is None:
            raise ImportValidationError("Choose an .xlsx file to upload.")
        name = (filename or "").lower()
        if name and not name.endswith(".xlsx"):
            raise ImportValidationError(
                "Please upload an Excel .xlsx file (not .xls or CSV)."
            )
        fh = file_obj  # type: ignore[assignment]

    try:
        result = _process_xlsx(fh, dry_run=False, skip_duplicates=skip_duplicates)
    finally:
        if cleanup_token:
            try:
                fh.close()
            except Exception:  # noqa: BLE001
                pass
            clear_staged(cleanup_token)
        elif hasattr(fh, "close") and staging_token is None and file_obj is None:
            pass

    return result


def _staging_dir() -> Path:
    root = Path(current_app.root_path) / "database" / ".import_staging"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _process_xlsx(
    file_obj: BinaryIO,
    *,
    dry_run: bool,
    skip_duplicates: bool,
) -> dict[str, Any]:
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ImportValidationError(
            "Could not read that Excel file. Re-save as .xlsx and try again."
        ) from exc

    sheet = wb.active
    if sheet is None:
        raise ImportValidationError("Workbook has no sheets.")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ImportValidationError("The sheet is empty.") from exc

    col_map = _map_headers(header_row or ())
    required = {"date", "amount", "description"}
    if not required.issubset(col_map):
        raise ImportValidationError(
            "Missing required columns. Need at least: date, amount, description, type, account."
        )
    # type / account optional — defaults applied in _row_to_payload

    lookups = _build_lookups()
    ready = 0
    duplicates = 0
    skipped = 0
    conflict_count = 0
    errors: list[str] = []
    preview_rows: list[dict[str, Any]] = []
    created_ids: list[int] = []

    for excel_row, values in enumerate(rows_iter, start=2):
        if _row_empty(values):
            continue
        # Skip leftover template placeholder rows with no amount
        amount_raw = _cell(values, col_map, "amount")
        desc_raw = str(_cell(values, col_map, "description") or "").strip()
        if (amount_raw is None or amount_raw == "") and not desc_raw:
            continue
        if desc_raw.endswith("— replace me") and (
            amount_raw in (None, "", 0, 0.0, "0", "500", 500)
        ):
            # still allow if user kept sample intentionally with real edits — only skip
            # untouched samples: description still ends with replace me AND amount is 500
            if amount_raw in (500, 500.0, "500") and "replace me" in desc_raw:
                skipped += 1
                preview_rows.append(
                    {
                        "row": excel_row,
                        "status": "skipped",
                        "summary": desc_raw,
                        "detail": "Template sample — replace or delete",
                        "envelope": "—",
                        "envelope_source": "",
                        "envelope_conflict": False,
                        "category_name": "",
                        "account_name": "",
                    }
                )
                continue

        try:
            payload = _row_to_payload(values, col_map, lookups)
            sheet_env = str(_cell(values, col_map, "envelope") or "").strip()
            meta = _preview_row_meta(payload, lookups, sheet_env_name=sheet_env)

            if skip_duplicates and _is_duplicate(payload):
                duplicates += 1
                preview_rows.append(
                    {
                        "row": excel_row,
                        "status": "duplicate",
                        "summary": payload["description"],
                        "detail": "Same date + amount + description already exists",
                        **meta,
                    }
                )
                continue

            if dry_run:
                ready += 1
                if meta.get("envelope_conflict"):
                    conflict_count += 1
                detail_parts = [
                    payload.get("transaction_type") or "",
                    meta.get("account_name") or "",
                ]
                if meta.get("category_name"):
                    detail_parts.append(meta["category_name"])
                if meta.get("envelope_conflict"):
                    detail_parts.append(
                        f"⚠ pot mismatch (category default {meta.get('default_envelope')})"
                    )
                elif meta.get("envelope") and meta["envelope"] != "—":
                    src = meta.get("envelope_source") or ""
                    detail_parts.append(
                        f"→ {meta['envelope']}"
                        + (f" ({src})" if src in ("sheet", "category") else "")
                    )
                preview_rows.append(
                    {
                        "row": excel_row,
                        "status": "ready",
                        "summary": (
                            f"{payload['date']} · {payload['amount']} · "
                            f"{payload['description']}"
                        ),
                        "detail": " · ".join(p for p in detail_parts if p),
                        **meta,
                    }
                )
            else:
                txn = create_transaction(payload)
                created_ids.append(txn.id)
                ready += 1
        except (ImportValidationError, TransactionValidationError, Exception) as exc:
            skipped += 1
            msg = str(exc)
            errors.append(f"Row {excel_row}: {msg}")
            preview_rows.append(
                {
                    "row": excel_row,
                    "status": "error",
                    "summary": desc_raw or f"Row {excel_row}",
                    "detail": msg,
                    "envelope": "—",
                    "envelope_source": "",
                    "envelope_conflict": False,
                    "category_name": "",
                    "account_name": "",
                }
            )

    return {
        "ready_count": ready,
        "created_count": 0 if dry_run else len(created_ids),
        "duplicate_count": duplicates,
        "skipped_count": skipped,
        "conflict_count": conflict_count,
        "error_count": len(errors),
        "errors": errors[:30],
        "preview_rows": preview_rows[:200],
        "created_ids": created_ids,
        "dry_run": dry_run,
    }


def _lookup_by_id(mapping: dict[str, Any], obj_id: int | None) -> Any:
    if not obj_id:
        return None
    for obj in mapping.values():
        if getattr(obj, "id", None) == obj_id:
            return obj
    return None


def _preview_row_meta(
    payload: dict[str, Any],
    lookups: dict[str, Any],
    *,
    sheet_env_name: str = "",
) -> dict[str, Any]:
    """Resolved envelope + category/envelope conflict for import preview."""
    account = _lookup_by_id(lookups["accounts"], payload.get("account_id"))
    category = _lookup_by_id(lookups["categories"], payload.get("category_id"))
    env = _lookup_by_id(lookups["envelopes"], payload.get("envelope_id"))
    default_env = None
    if category and category.envelope_id:
        default_env = _lookup_by_id(lookups["envelopes"], category.envelope_id)

    is_joint = False
    if account:
        is_joint = (account.owner or "").lower() == "joint" or account.name == "Joint Account"

    if env:
        if sheet_env_name:
            source = "sheet"
        elif default_env and default_env.id == env.id:
            source = "category"
        else:
            source = "resolved"
        envelope_name = env.name
    elif not is_joint and payload.get("transaction_type") == "expense":
        source = "n/a"
        envelope_name = "— (personal account)"
    else:
        source = ""
        envelope_name = "—"

    conflict = bool(
        payload.get("transaction_type") == "expense"
        and is_joint
        and env
        and default_env
        and env.id != default_env.id
    )

    return {
        "envelope": envelope_name,
        "envelope_source": source,
        "envelope_conflict": conflict,
        "default_envelope": default_env.name if default_env else "",
        "category_name": category.name if category else "",
        "account_name": account.name if account else "",
    }


def _is_duplicate(payload: dict[str, Any]) -> bool:
    from models import Transaction as Txn

    return (
        Txn.query.filter(
            Txn.date == date.fromisoformat(payload["date"]),
            Txn.amount == Decimal(payload["amount"]),
            Txn.description == payload["description"],
        ).first()
        is not None
    )


def _norm_header(value: Any) -> str:
    text = re.sub(r"[\s_]+", "_", str(value or "").strip().lower())
    aliases = {
        "from_account": "account",
        "account_name": "account",
        "to": "to_account",
        "destination": "to_account",
        "envelope_name": "envelope",
        "txn_type": "type",
        "transaction_type": "type",
    }
    return aliases.get(text, text)


def _map_headers(header_row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _row_empty(values: tuple | None) -> bool:
    if not values:
        return True
    return all(v is None or str(v).strip() == "" for v in values)


def _cell(values: tuple, col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def _build_lookups() -> dict[str, Any]:
    accounts = {
        _key(a.name): a
        for a in Account.query.filter_by(is_active=True).all()
    }
    categories = {
        _key(c.name): c
        for c in Category.query.filter_by(is_active=True, parent_id=None).all()
    }
    envelopes = {
        _key(e.name): e
        for e in Envelope.query.filter_by(is_active=True).all()
    }
    investments = {
        _key(i.name): i
        for i in Investment.query.filter_by(is_active=True).all()
    }
    return {
        "accounts": accounts,
        "categories": categories,
        "envelopes": envelopes,
        "investments": investments,
        "joint": _joint_account(),
    }


def _key(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _parse_excel_date(value: Any) -> str:
    if value is None or value == "":
        raise ImportValidationError("date is required.")
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    raise ImportValidationError(f"Unrecognized date “{value}”. Use YYYY-MM-DD.")


def _parse_excel_amount(value: Any) -> str:
    if value is None or value == "":
        raise ImportValidationError("amount is required.")
    if isinstance(value, (int, float, Decimal)):
        amount = Decimal(str(value))
    else:
        text = str(value).strip().replace(",", "").replace("₹", "")
        amount = Decimal(text)
    if amount <= 0:
        raise ImportValidationError("amount must be greater than zero.")
    return str(amount.quantize(Decimal("0.01")))


def _row_to_payload(
    values: tuple, col_map: dict[str, int], lookups: dict[str, Any]
) -> dict[str, Any]:
    txn_type = str(_cell(values, col_map, "type") or "expense").strip().lower()
    if txn_type not in Transaction.TRANSACTION_TYPES:
        raise ImportValidationError(
            f"Invalid type “{txn_type}”. Use expense/income/transfer/investment/refund."
        )

    account_name = str(_cell(values, col_map, "account") or "").strip()
    account = None
    if account_name:
        account = lookups["accounts"].get(_key(account_name))
        if not account:
            raise ImportValidationError(
                f"Unknown account “{account_name}”. Check the Reference sheet."
            )
    elif txn_type in ("expense", "income", "refund", "investment"):
        account = lookups.get("joint")
        if not account:
            raise ImportValidationError(
                "account is empty — set Joint Account in the app, or fill the account column."
            )
    else:
        raise ImportValidationError("account is required for transfers.")

    payload: dict[str, Any] = {
        "date": _parse_excel_date(_cell(values, col_map, "date")),
        "amount": _parse_excel_amount(_cell(values, col_map, "amount")),
        "description": str(_cell(values, col_map, "description") or "").strip(),
        "transaction_type": txn_type,
        "account_id": account.id,
        "notes": str(_cell(values, col_map, "notes") or "").strip() or None,
    }
    if not payload["description"]:
        raise ImportValidationError("description is required.")

    to_name = str(_cell(values, col_map, "to_account") or "").strip()
    if txn_type == "transfer":
        if not to_name:
            raise ImportValidationError("to_account is required for transfers.")
        to_acc = lookups["accounts"].get(_key(to_name))
        if not to_acc:
            raise ImportValidationError(f"Unknown to_account “{to_name}”.")
        payload["to_account_id"] = to_acc.id

    cat_name = str(_cell(values, col_map, "category") or "").strip()
    category = None
    if cat_name:
        category = lookups["categories"].get(_key(cat_name))
        if not category:
            raise ImportValidationError(f"Unknown category “{cat_name}”.")
        payload["category_id"] = category.id

    # Envelopes only apply when spending from Joint
    is_joint = (account.owner or "").lower() == "joint" or account.name == "Joint Account"
    env_name = str(_cell(values, col_map, "envelope") or "").strip()
    if is_joint:
        if env_name:
            env = lookups["envelopes"].get(_key(env_name))
            if not env:
                raise ImportValidationError(f"Unknown envelope “{env_name}”.")
            payload["envelope_id"] = env.id
        elif category and category.envelope_id:
            payload["envelope_id"] = category.envelope_id

    inv_name = str(_cell(values, col_map, "investment") or "").strip()
    if inv_name and txn_type == "investment":
        inv = lookups["investments"].get(_key(inv_name))
        if not inv:
            raise ImportValidationError(f"Unknown investment “{inv_name}”.")
        payload["investment_id"] = inv.id

    paid_by = str(_cell(values, col_map, "paid_by") or "").strip().lower()
    if paid_by:
        payload["paid_by"] = paid_by

    payment_mode = str(_cell(values, col_map, "payment_mode") or "upi").strip().lower()
    if payment_mode:
        payload["payment_mode"] = payment_mode.replace(" ", "_")

    need_want = str(_cell(values, col_map, "need_want") or "").strip().lower()
    if need_want:
        payload["need_want"] = need_want

    return payload
